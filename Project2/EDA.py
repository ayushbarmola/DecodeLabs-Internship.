
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# 1. FILE PATHS
# ------------------------------------------------------------

input_file = r"C:\Users\ASUS\Desktop\Data_Analytics\Exploratory Data Analysis\Cleaned_Dataset.xlsx"
output_file = "C:\Users\ASUS\Desktop\Data_Analytics\Exploratory Data Analysis\Dataset_Analysis_Report.xlsx"


# ------------------------------------------------------------
# 2. READ DATA
# ------------------------------------------------------------

df = pd.read_excel(input_file)

print("Dataset loaded successfully!")
print("Rows:", len(df))
print("Columns:", len(df.columns))


# ------------------------------------------------------------
# 3. BASIC DATA PREPARATION
# ------------------------------------------------------------

df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

numeric_columns = [
    "Quantity",
    "UnitPrice",
    "ItemsInCart",
    "TotalPrice"
]

for col in numeric_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce")


# ------------------------------------------------------------
# 4. BASIC STATISTICS
# ------------------------------------------------------------

statistics = []

for col in numeric_columns:

    statistics.append({
        "Variable": col,
        "Count": df[col].count(),
        "Mean": df[col].mean(),
        "Median": df[col].median(),
        "Minimum": df[col].min(),
        "Maximum": df[col].max(),
        "Standard Deviation": df[col].std()
    })

stats_df = pd.DataFrame(statistics)


# ------------------------------------------------------------
# 5. YEARLY SALES TREND
# ------------------------------------------------------------

df["Year"] = df["Date"].dt.year

yearly_sales = (
    df.groupby("Year")
      .agg(
          Orders=("OrderID", "count"),
          Revenue=("TotalPrice", "sum"),
          Average_Order_Value=("TotalPrice", "mean")
      )
      .reset_index()
)

# Round values
yearly_sales["Revenue"] = yearly_sales["Revenue"].round(2)
yearly_sales["Average_Order_Value"] = yearly_sales[
    "Average_Order_Value"
].round(2)


# ------------------------------------------------------------
# 6. PRODUCT ANALYSIS
# ------------------------------------------------------------

product_analysis = (
    df.groupby("Product")
      .agg(
          Orders=("OrderID", "count"),
          Quantity_Sold=("Quantity", "sum"),
          Revenue=("TotalPrice", "sum"),
          Average_Order_Value=("TotalPrice", "mean")
      )
      .reset_index()
      .sort_values("Revenue", ascending=False)
)

product_analysis[
    ["Revenue", "Average_Order_Value"]
] = product_analysis[
    ["Revenue", "Average_Order_Value"]
].round(2)


# ------------------------------------------------------------
# 7. PAYMENT METHOD ANALYSIS
# ------------------------------------------------------------

payment_analysis = (
    df.groupby("PaymentMethod")
      .agg(
          Orders=("OrderID", "count"),
          Revenue=("TotalPrice", "sum"),
          Average_Order_Value=("TotalPrice", "mean")
      )
      .reset_index()
      .sort_values("Orders", ascending=False)
)

payment_analysis[
    ["Revenue", "Average_Order_Value"]
] = payment_analysis[
    ["Revenue", "Average_Order_Value"]
].round(2)


# ------------------------------------------------------------
# 8. ORDER STATUS ANALYSIS
# ------------------------------------------------------------

status_analysis = (
    df.groupby("OrderStatus")
      .agg(
          Orders=("OrderID", "count"),
          Revenue=("TotalPrice", "sum"),
          Average_Order_Value=("TotalPrice", "mean")
      )
      .reset_index()
      .sort_values("Orders", ascending=False)
)

status_analysis[
    ["Revenue", "Average_Order_Value"]
] = status_analysis[
    ["Revenue", "Average_Order_Value"]
].round(2)

status_analysis["Percentage"] = (
    status_analysis["Orders"] / len(df) * 100
).round(2)


# ------------------------------------------------------------
# 9. REFERRAL SOURCE ANALYSIS
# ------------------------------------------------------------

referral_analysis = (
    df.groupby("ReferralSource")
      .agg(
          Orders=("OrderID", "count"),
          Revenue=("TotalPrice", "sum"),
          Average_Order_Value=("TotalPrice", "mean")
      )
      .reset_index()
      .sort_values("Orders", ascending=False)
)

referral_analysis[
    ["Revenue", "Average_Order_Value"]
] = referral_analysis[
    ["Revenue", "Average_Order_Value"]
].round(2)


# ------------------------------------------------------------
# 10. OUTLIER ANALYSIS USING IQR
# ------------------------------------------------------------

outlier_summary = []

outlier_records = []

for col in numeric_columns:

    Q1 = df[col].quantile(0.25)
    Q3 = df[col].quantile(0.75)

    IQR = Q3 - Q1

    lower_limit = Q1 - 1.5 * IQR
    upper_limit = Q3 + 1.5 * IQR

    outliers = df[
        (df[col] < lower_limit) |
        (df[col] > upper_limit)
    ]

    outlier_summary.append({
        "Variable": col,
        "Q1": round(Q1, 2),
        "Q3": round(Q3, 2),
        "IQR": round(IQR, 2),
        "Lower Limit": round(lower_limit, 2),
        "Upper Limit": round(upper_limit, 2),
        "Outlier Count": len(outliers)
    })

    if len(outliers) > 0:

        temp = outliers.copy()

        temp["Outlier_Variable"] = col
        temp["Lower_Limit"] = lower_limit
        temp["Upper_Limit"] = upper_limit

        outlier_records.append(temp)


outlier_summary_df = pd.DataFrame(outlier_summary)

if outlier_records:
    outliers_df = pd.concat(
        outlier_records,
        ignore_index=True
    )
else:
    outliers_df = pd.DataFrame()


# ------------------------------------------------------------
# 11. GENERATE KEY OBSERVATIONS
# ------------------------------------------------------------

observations = []


# Mean vs Median
total_mean = df["TotalPrice"].mean()
total_median = df["TotalPrice"].median()

if total_mean > total_median:
    observations.append(
        "Total Price has a higher mean than median, "
        "indicating a right-skewed distribution caused by "
        "some high-value orders."
    )
elif total_mean < total_median:
    observations.append(
        "Total Price has a lower mean than median, "
        "indicating a left-skewed distribution."
    )
else:
    observations.append(
        "Mean and median Total Price are approximately equal."
    )


# Average quantity
observations.append(
    f"The average order contains "
    f"{df['Quantity'].mean():.2f} items."
)


# Product
best_product = product_analysis.iloc[0]["Product"]
best_product_revenue = product_analysis.iloc[0]["Revenue"]

worst_product = product_analysis.iloc[-1]["Product"]
worst_product_revenue = product_analysis.iloc[-1]["Revenue"]

observations.append(
    f"{best_product} generated the highest revenue "
    f"of ₹{best_product_revenue:,.2f}."
)

observations.append(
    f"{worst_product} generated the lowest revenue "
    f"of ₹{worst_product_revenue:,.2f}."
)


# Payment
best_payment = payment_analysis.iloc[0]["PaymentMethod"]

observations.append(
    f"{best_payment} is the most frequently used payment method."
)


# Referral
best_referral = referral_analysis.iloc[0]["ReferralSource"]

observations.append(
    f"{best_referral} generated the highest number of orders."
)


# Order status
cancelled = df["OrderStatus"].eq("Cancelled").sum()
returned = df["OrderStatus"].eq("Returned").sum()

cancel_return_percentage = (
    (cancelled + returned) / len(df) * 100
)

observations.append(
    f"Cancelled and Returned orders together account for "
    f"{cancel_return_percentage:.2f}% of all orders."
)


# Outliers
total_outliers = len(
    outlier_summary_df[
        outlier_summary_df["Outlier Count"] > 0
    ]
)

observations.append(
    f"{total_outliers} numeric variables contain potential outliers "
    f"according to the IQR method."
)


# Create observation dataframe
observations_df = pd.DataFrame({
    "Key Observation": observations
})


# ------------------------------------------------------------
# 12. CREATE EXCEL FILE
# ------------------------------------------------------------

with pd.ExcelWriter(
    output_file,
    engine="openpyxl"
) as writer:

    # Original cleaned data
    df.to_excel(
        writer,
        sheet_name="Cleaned Data",
        index=False
    )

    # Analysis sheets
    stats_df.to_excel(
        writer,
        sheet_name="Basic Statistics",
        index=False
    )

    yearly_sales.to_excel(
        writer,
        sheet_name="Yearly Trend",
        index=False
    )

    product_analysis.to_excel(
        writer,
        sheet_name="Product Analysis",
        index=False
    )

    payment_analysis.to_excel(
        writer,
        sheet_name="Payment Analysis",
        index=False
    )

    status_analysis.to_excel(
        writer,
        sheet_name="Order Status",
        index=False
    )

    referral_analysis.to_excel(
        writer,
        sheet_name="Referral Analysis",
        index=False
    )

    outlier_summary_df.to_excel(
        writer,
        sheet_name="Outlier Summary",
        index=False
    )

    if not outliers_df.empty:
        outliers_df.to_excel(
            writer,
            sheet_name="Outlier Records",
            index=False
        )

    observations_df.to_excel(
        writer,
        sheet_name="Key Observations",
        index=False
    )


# ------------------------------------------------------------
# 13. FORMAT EXCEL WORKBOOK
# ------------------------------------------------------------

wb = load_workbook(output_file)

# Styles
header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

title_font = Font(
    bold=True,
    size=14
)

thin_border = Border(
    bottom=Side(
        style="thin",
        color="D9E1F2"
    )
)


# Format every worksheet
for ws in wb.worksheets:

    # Header formatting
    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto column width
    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            try:
                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 45)


# ------------------------------------------------------------
# 14. NUMBER FORMATTING
# ------------------------------------------------------------

for ws_name in [
    "Basic Statistics",
    "Yearly Trend",
    "Product Analysis",
    "Payment Analysis",
    "Order Status",
    "Referral Analysis",
    "Outlier Summary"
]:

    ws = wb[ws_name]

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell.value, float):

                cell.number_format = '#,##0.00'


# Currency formatting
currency_sheets = [
    "Yearly Trend",
    "Product Analysis",
    "Payment Analysis",
    "Order Status",
    "Referral Analysis"
]

for sheet_name in currency_sheets:

    ws = wb[sheet_name]

    for row in ws.iter_rows():

        for cell in row:

            if cell.column >= 1:

                header = ws.cell(
                    row=1,
                    column=cell.column
                ).value

                if header in [
                    "Revenue",
                    "Average_Order_Value"
                ]:

                    cell.number_format = '₹#,##0.00'


# ------------------------------------------------------------
# 15. CREATE CHARTS
# ------------------------------------------------------------

# Yearly revenue chart
ws = wb["Yearly Trend"]

chart = LineChart()

chart.title = "Yearly Revenue Trend"
chart.y_axis.title = "Revenue"
chart.x_axis.title = "Year"

data = Reference(
    ws,
    min_col=3,
    min_row=1,
    max_row=ws.max_row
)

categories = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)

chart.add_data(
    data,
    titles_from_data=True
)

chart.set_categories(categories)

chart.height = 8
chart.width = 15

ws.add_chart(chart, "F2")


# Product revenue chart
ws = wb["Product Analysis"]

chart = BarChart()

chart.title = "Revenue by Product"
chart.y_axis.title = "Revenue"
chart.x_axis.title = "Product"

data = Reference(
    ws,
    min_col=4,
    min_row=1,
    max_row=ws.max_row
)

categories = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)

chart.add_data(
    data,
    titles_from_data=True
)

chart.set_categories(categories)

chart.height = 8
chart.width = 15

ws.add_chart(chart, "H2")


# Order status chart
ws = wb["Order Status"]

chart = BarChart()

chart.title = "Orders by Status"
chart.y_axis.title = "Number of Orders"
chart.x_axis.title = "Status"

data = Reference(
    ws,
    min_col=2,
    min_row=1,
    max_row=ws.max_row
)

categories = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)

chart.add_data(
    data,
    titles_from_data=True
)

chart.set_categories(categories)

chart.height = 8
chart.width = 15

ws.add_chart(chart, "H2")


# ------------------------------------------------------------
# 16. SAVE WORKBOOK
# ------------------------------------------------------------

wb.save(output_file)

print("\n==========================================")
print("DATA ANALYSIS COMPLETED SUCCESSFULLY!")
print("==========================================")
print(f"Output file: {output_file}")

print("\nSheets created:")

for sheet in wb.sheetnames:
    print(" -", sheet)